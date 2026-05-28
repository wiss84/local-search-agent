"""
XLSX parser for the Local Search Agent ingestion pipeline.

Uses openpyxl for Excel workbook extraction.

Strategy:
- Each sheet becomes a Markdown section (## Sheet Name)
- Each sheet's data is rendered as a Markdown table
- Empty sheets are skipped
- Cell values are stringified cleanly (dates, numbers, booleans)
- Formula results (cached values) are used, not formula strings

Install: pip install "openpyxl>=3.1.0"
"""

from __future__ import annotations

import logging
import os
from datetime import date, datetime
from typing import Any, Optional

from local_search_agent.core.document_node import DocumentNode
from local_search_agent.ingestion.cleaner import clean
from local_search_agent.ingestion.parser import BaseParser, ParserError

logger = logging.getLogger(__name__)


def _cell_value(value: Any) -> str:
    """Convert a cell value to a clean string representation."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        # Avoid "1.0" for whole numbers stored as floats
        if value == int(value):
            return str(int(value))
        return f"{value:.4g}"
    return str(value).strip()


def _sheet_to_markdown(sheet) -> str:
    """
    Convert a single openpyxl worksheet to a Markdown table.

    Reads only the populated data region (min_row to max_row).
    Returns empty string if sheet has no data.
    """
    rows = list(sheet.iter_rows(values_only=True))

    # Filter out completely empty rows
    rows = [r for r in rows if any(c is not None for c in r)]
    if not rows:
        return ""

    # Normalise row widths
    max_cols = max(len(r) for r in rows)
    padded = [list(r) + [None] * (max_cols - len(r)) for r in rows]

    str_rows = [[_cell_value(cell) for cell in row] for row in padded]

    # First row → header
    header = str_rows[0]
    separator = ["---"] * max_cols
    body = str_rows[1:]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in body:
        escaped = [cell.replace("|", "\\|") for cell in row]
        lines.append("| " + " | ".join(escaped) + " |")

    return "\n".join(lines)


class XLSXParser(BaseParser):
    """
    Parse Excel workbooks (.xlsx, .xlsm) using openpyxl.

    Each visible worksheet becomes a ## section with a Markdown table.
    Hidden sheets are skipped by default (can be overridden).
    """

    def __init__(self, include_hidden_sheets: bool = False):
        self._include_hidden = include_hidden_sheets

    @property
    def supported_extensions(self) -> frozenset[str]:
        return frozenset({".xlsx", ".xlsm", ".xls"})

    def parse(
        self,
        source_path: str,
        workspace: str,
        title: Optional[str] = None,
    ) -> DocumentNode:
        if not os.path.isfile(source_path):
            raise FileNotFoundError(f"XLSX file not found: {source_path!r}")

        try:
            import openpyxl
        except ImportError as e:
            raise ParserError(
                source_path,
                "openpyxl is not installed. Run: pip install 'openpyxl>=3.1.0'",
                original=e,
            )

        logger.info("Parsing XLSX: %s", source_path)

        try:
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        except Exception as e:
            raise ParserError(source_path, f"openpyxl failed to open workbook: {e}", original=e)

        sections: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Skip hidden sheets unless requested
                if not self._include_hidden and ws.sheet_state != "visible":
                    logger.debug("Skipping hidden sheet %r in %s", sheet_name, source_path)
                    continue

                md_table = _sheet_to_markdown(ws)
                if not md_table:
                    logger.debug("Skipping empty sheet %r in %s", sheet_name, source_path)
                    continue

                sections.append(f"## {sheet_name}\n\n{md_table}")
        finally:
            wb.close()

        if not sections:
            raise ParserError(
                source_path, "Workbook contains no readable data in any visible sheet."
            )

        raw_text = "\n\n".join(sections)
        cleaned_text = clean(raw_text)

        return DocumentNode.from_file(
            source_path=source_path,
            text=cleaned_text,
            workspace=workspace,
            title=title,
        )
