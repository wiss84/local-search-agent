"""
Excel (.xlsx) export for a single table extracted from a chat answer.

The frontend reads one rendered <table> element straight out of the DOM
(headers + rows as plain strings) and posts it here — no Markdown
re-parsing involved, since by the time a table reaches the chat it has
already been rendered to real HTML by marked.js.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font


def build_xlsx(headers: list[str], rows: list[list[str]]) -> bytes:
    """Build a single-sheet .xlsx workbook from a header row + data rows, return raw bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    # Roughly auto-size columns based on content length (capped to avoid runaway widths)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
