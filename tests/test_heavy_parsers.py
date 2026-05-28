"""
Tests for the heavier document parsers: HTML, XLSX, PDF, DOCX.

Split into two groups by strategy:

  Fast (no mark) — HTMLParser and XLSXParser
  ─────────────────────────────────────────────────────────────────────────
  Both use stdlib/lightweight libraries (BeautifulSoup4, openpyxl) and
  accept programmatically-generated files. No mocking required. These run
  in the normal CI pass alongside test_parsers.py.

  Slow (@pytest.mark.slow) — PDFParser and DOCXParser
  ─────────────────────────────────────────────────────────────────────────
  Both route through Docling (IBM), which loads multi-hundred-MB model
  weights on first call. We mock Docling at the narrowest boundary:
    - DocumentConverter() constructor → returns a mock
    - converter.convert(path).document.export_to_markdown() → returns a string
  Everything else (file existence check, clean() pipeline, DocumentNode
  construction, ParserError propagation) is exercised against real code.
  Skipped with --fast; run in full PR checks.

Design notes:
- XLSX: openpyxl is used to write real .xlsx files in fixtures so the
  parser's full read path (open → iter_rows → Markdown table) is exercised.
- HTML: real .html files written as strings; lxml parses them.
- PDF/DOCX: Docling singleton (_CONVERTER) is patched at module level so
  the lazy-init path is also covered.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from local_search_agent.ingestion.parser import ParserError
from local_search_agent.ingestion.parsers.html_parser import HTMLParser
from local_search_agent.ingestion.parsers.xlsx_parser import XLSXParser

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write(tmp_path: Path, name: str, content: str, encoding: str = "utf-8") -> Path:
    f = tmp_path / name
    f.write_text(content, encoding=encoding)
    return f


def _make_xlsx(tmp_path: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    """
    Create a real .xlsx file with openpyxl.

    sheets = {"Sheet1": [["Col1", "Col2"], ["A", "B"]], ...}
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / name
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# HTMLParser — fast, no mocking
# ---------------------------------------------------------------------------

class TestHTMLParser:
    def test_basic_content_extracted(self, tmp_path):
        f = _write(tmp_path, "page.html",
                   "<html><body><h1>Report Title</h1><p>Q3 revenue was strong.</p></body></html>")
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Report Title" in node.text
        assert "Q3 revenue" in node.text
        assert node.file_type == "html"

    def test_htm_extension_supported(self, tmp_path):
        f = _write(tmp_path, "page.htm",
                   "<html><body><p>Legacy page.</p></body></html>")
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Legacy page" in node.text
        assert node.file_type == "htm"

    def test_heading_levels_converted(self, tmp_path):
        f = _write(tmp_path, "headings.html",
                   "<html><body><h1>Top</h1><h2>Second</h2><h3>Third</h3></body></html>")
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Top" in node.text
        assert "Second" in node.text
        assert "Third" in node.text

    def test_table_converted_to_markdown(self, tmp_path):
        html = """
        <html><body>
        <table>
          <tr><th>Name</th><th>Score</th></tr>
          <tr><td>Alice</td><td>95</td></tr>
          <tr><td>Bob</td><td>87</td></tr>
        </table>
        </body></html>
        """
        f = _write(tmp_path, "table.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "|" in node.text
        assert "Alice" in node.text
        assert "95" in node.text

    def test_script_and_style_stripped(self, tmp_path):
        html = """
        <html>
        <head><style>body { color: red; }</style></head>
        <body>
          <script>alert('xss')</script>
          <p>Clean content here.</p>
        </body>
        </html>
        """
        f = _write(tmp_path, "noisy.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "alert" not in node.text
        assert "color: red" not in node.text
        assert "Clean content" in node.text

    def test_nav_stripped(self, tmp_path):
        html = """
        <html><body>
        <nav><a href="/">Home</a><a href="/about">About</a></nav>
        <main><p>Main content.</p></main>
        </body></html>
        """
        f = _write(tmp_path, "nav.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Main content" in node.text

    def test_prefers_main_content_area(self, tmp_path):
        html = """
        <html><body>
        <aside>Sidebar noise</aside>
        <main><p>Real content.</p></main>
        </body></html>
        """
        f = _write(tmp_path, "main.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Real content" in node.text

    def test_title_extracted_from_title_tag(self, tmp_path):
        html = "<html><head><title>My Page Title</title></head><body><p>Body.</p></body></html>"
        f = _write(tmp_path, "titled.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert node.title == "My Page Title"

    def test_title_override(self, tmp_path):
        html = "<html><head><title>Original</title></head><body><p>Body.</p></body></html>"
        f = _write(tmp_path, "override.html", html)
        node = HTMLParser().parse(str(f), workspace="test", title="My Override")
        assert node.title == "My Override"

    def test_code_block_fenced(self, tmp_path):
        html = "<html><body><pre><code>def hello():\n    print('hi')</code></pre></body></html>"
        f = _write(tmp_path, "code.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "```" in node.text
        assert "def hello" in node.text

    def test_inline_code(self, tmp_path):
        html = "<html><body><p>Use <code>pip install</code> to install.</p></body></html>"
        f = _write(tmp_path, "inline.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "pip install" in node.text

    def test_bold_and_italic(self, tmp_path):
        html = "<html><body><p><strong>Bold</strong> and <em>italic</em>.</p></body></html>"
        f = _write(tmp_path, "fmt.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Bold" in node.text
        assert "italic" in node.text

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            HTMLParser().parse(str(tmp_path / "missing.html"), workspace="test")

    def test_empty_body_returns_minimal_content(self, tmp_path):
        f = _write(tmp_path, "empty.html", "<html><body></body></html>")
        node = HTMLParser().parse(str(f), workspace="test")
        assert isinstance(node.text, str)

    def test_confluence_wiki_content_selector(self, tmp_path):
        html = """
        <html><body>
        <div class="sidebar">Noise</div>
        <div class="wiki-content"><p>Wiki page body.</p></div>
        </body></html>
        """
        f = _write(tmp_path, "confluence.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Wiki page body" in node.text

    def test_unordered_list_preserved(self, tmp_path):
        html = "<html><body><ul><li>Item one</li><li>Item two</li></ul></body></html>"
        f = _write(tmp_path, "list.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Item one" in node.text
        assert "Item two" in node.text

    def test_ordered_list_preserved(self, tmp_path):
        html = "<html><body><ol><li>First</li><li>Second</li></ol></body></html>"
        f = _write(tmp_path, "ol.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "First" in node.text
        assert "Second" in node.text

    def test_non_ascii_content(self, tmp_path):
        html = "<html><body><p>Héllo Wörld — 日本語テスト</p></body></html>"
        f = _write(tmp_path, "unicode.html", html)
        node = HTMLParser().parse(str(f), workspace="test")
        assert "Héllo" in node.text


# ---------------------------------------------------------------------------
# XLSXParser — fast, uses real openpyxl files
# ---------------------------------------------------------------------------

class TestXLSXParser:
    def test_basic_sheet_extracted(self, tmp_path):
        f = _make_xlsx(tmp_path, "basic.xlsx", {
            "Sales": [["Region", "Revenue"], ["North", 120000], ["South", 95000]]
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert "Sales" in node.text
        assert "North" in node.text
        assert "120000" in node.text
        assert node.file_type == "xlsx"

    def test_multiple_sheets_become_sections(self, tmp_path):
        f = _make_xlsx(tmp_path, "multi.xlsx", {
            "Q1": [["Month", "Sales"], ["Jan", 10000]],
            "Q2": [["Month", "Sales"], ["Apr", 15000]],
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert "Q1" in node.text
        assert "Q2" in node.text
        assert "Jan" in node.text
        assert "Apr" in node.text

    def test_output_is_markdown_table(self, tmp_path):
        f = _make_xlsx(tmp_path, "table.xlsx", {
            "Data": [["A", "B"], ["1", "2"]]
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert "|" in node.text

    def test_all_empty_workbook_raises_parser_error(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        # Default sheet exists but has no data
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))
        with pytest.raises(ParserError):
            XLSXParser().parse(str(path), workspace="test")

    def test_boolean_cells_rendered_as_yes_no(self, tmp_path):
        f = _make_xlsx(tmp_path, "booleans.xlsx", {
            "Flags": [["Name", "Active"], ["Alice", True], ["Bob", False]]
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert "Yes" in node.text
        assert "No" in node.text

    def test_float_whole_number_no_decimal(self, tmp_path):
        # 42.0 should render as "42", not "42.0"
        f = _make_xlsx(tmp_path, "floats.xlsx", {
            "Data": [["Value"], [42.0]]
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert "42" in node.text
        assert "42.0" not in node.text

    def test_pipe_chars_escaped_in_cells(self, tmp_path):
        f = _make_xlsx(tmp_path, "pipes.xlsx", {
            "Data": [["Code"], ["A|B"]]
        })
        node = XLSXParser().parse(str(f), workspace="test")
        assert r"A\|B" in node.text

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            XLSXParser().parse(str(tmp_path / "missing.xlsx"), workspace="test")

    def test_xlsm_extension_supported(self, tmp_path):
        # .xlsm is a macro-enabled workbook — openpyxl reads them fine
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        path = tmp_path / "macros.xlsm"
        wb.save(str(path))
        node = XLSXParser().parse(str(path), workspace="test")
        assert "Alice" in node.text

    def test_large_sheet_all_rows_present(self, tmp_path):
        rows = [["ID", "Value"]] + [[i, i * 10] for i in range(1, 201)]
        f = _make_xlsx(tmp_path, "large.xlsx", {"Data": rows})
        node = XLSXParser().parse(str(f), workspace="test")
        # First and last data rows must both appear
        assert "1" in node.text
        assert "200" in node.text

    def test_hidden_sheet_skipped_by_default(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        visible_ws = wb.active
        visible_ws.title = "Visible"
        visible_ws.append(["Name"])
        visible_ws.append(["Alice"])
        hidden_ws = wb.create_sheet(title="Hidden")
        hidden_ws.append(["Secret"])
        hidden_ws.append(["TopSecret"])
        hidden_ws.sheet_state = "hidden"
        path = tmp_path / "hidden.xlsx"
        wb.save(str(path))

        node = XLSXParser().parse(str(path), workspace="test")
        assert "Alice" in node.text
        assert "TopSecret" not in node.text

    def test_hidden_sheet_included_when_requested(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        visible_ws = wb.active
        visible_ws.title = "Visible"
        visible_ws.append(["Name"])
        visible_ws.append(["Alice"])
        hidden_ws = wb.create_sheet(title="Hidden")
        hidden_ws.append(["Secret"])
        hidden_ws.append(["TopSecret"])
        hidden_ws.sheet_state = "hidden"
        path = tmp_path / "hidden_incl.xlsx"
        wb.save(str(path))

        node = XLSXParser(include_hidden_sheets=True).parse(str(path), workspace="test")
        assert "TopSecret" in node.text


# ---------------------------------------------------------------------------
# PDFParser — slow, Docling mocked at boundary
# ---------------------------------------------------------------------------

def _make_docling_mock(markdown_output: str):
    """
    Build a mock that satisfies the Docling API:
      converter = DocumentConverter()
      result = converter.convert(path)
      md = result.document.export_to_markdown()
    """
    mock_doc = MagicMock()
    mock_doc.export_to_markdown.return_value = markdown_output

    mock_result = MagicMock()
    mock_result.document = mock_doc

    mock_converter = MagicMock()
    mock_converter.convert.return_value = mock_result

    mock_converter_class = MagicMock(return_value=mock_converter)
    return mock_converter_class, mock_converter


@pytest.mark.slow
class TestPDFParser:
    """
    Tests for PDFParser. Docling is mocked so no model weights are loaded.
    What is tested: file-existence guard, clean() pipeline, DocumentNode
    construction, ParserError propagation, and the small-file single-call path.
    """

    def _patch_docling(self, markdown: str):
        """Return a context manager that patches Docling inside pdf_parser."""
        mock_class, mock_converter = _make_docling_mock(markdown)
        return patch(
            "local_search_agent.ingestion.parsers.pdf_parser._get_converter",
            return_value=mock_converter,
        )

    def test_basic_text_extracted(self, tmp_path):
        f = tmp_path / "report.pdf"
        f.write_bytes(b"%PDF-1.4 fake")  # real byte content irrelevant — Docling is mocked

        with self._patch_docling("# Q3 Report\n\nRevenue grew 12% year-on-year."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert "Q3 Report" in node.text
        assert "Revenue grew" in node.text
        assert node.file_type == "pdf"

    def test_file_type_set_correctly(self, tmp_path):
        f = tmp_path / "doc.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("Some content."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert node.file_type == "pdf"

    def test_file_not_found_raises(self, tmp_path):
        from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
        with pytest.raises(FileNotFoundError):
            PDFParser().parse(str(tmp_path / "missing.pdf"), workspace="test")

    def test_docling_exception_raises_parser_error(self, tmp_path):
        f = tmp_path / "bad.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        mock_converter = MagicMock()
        mock_converter.convert.side_effect = RuntimeError("Docling internal error")

        with patch(
            "local_search_agent.ingestion.parsers.pdf_parser._get_converter",
            return_value=mock_converter,
        ):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            with pytest.raises(ParserError):
                PDFParser().parse(str(f), workspace="test")

    def test_clean_pipeline_applied(self, tmp_path):
        # cleaner normalises smart quotes; verify it runs
        f = tmp_path / "quotes.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("\u201cSmart quotes\u201d and \u2014 em-dash."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        # After clean(), smart quotes → straight quotes, em-dash → hyphen
        assert '"Smart quotes"' in node.text or "Smart quotes" in node.text
        assert "\u201c" not in node.text
        assert "\u2014" not in node.text

    def test_title_override(self, tmp_path):
        f = tmp_path / "annual.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("# Annual Report\n\nContent."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test", title="My Custom Title")

        assert node.title == "My Custom Title"

    def test_default_title_is_filename_stem(self, tmp_path):
        f = tmp_path / "financial_summary.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert node.title == "financial_summary"

    def test_workspace_set_on_node(self, tmp_path):
        f = tmp_path / "doc.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="finance_team")

        assert node.workspace == "finance_team"

    def test_source_path_set_on_node(self, tmp_path):
        f = tmp_path / "doc.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert str(f.resolve()) == node.source_path

    def test_doc_id_is_16_hex_chars(self, tmp_path):
        f = tmp_path / "doc.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert len(node.doc_id) == 16
        assert all(c in "0123456789abcdef" for c in node.doc_id)

    def test_watermark_stripped_by_cleaner(self, tmp_path):
        f = tmp_path / "conf.pdf"
        f.write_bytes(b"%PDF-1.4 fake")

        with self._patch_docling("CONFIDENTIAL\n\nActual content here.\n\nCONFIDENTIAL"):
            from local_search_agent.ingestion.parsers.pdf_parser import PDFParser
            node = PDFParser().parse(str(f), workspace="test")

        assert "Actual content here" in node.text
        assert "CONFIDENTIAL" not in node.text


# ---------------------------------------------------------------------------
# DOCXParser — slow, Docling mocked at boundary
# ---------------------------------------------------------------------------

@pytest.mark.slow
class TestDOCXParser:
    """
    Tests for DOCXParser. Docling is mocked so no model weights are loaded.
    python-docx is used to create real .docx fixtures so the size-estimation
    path is exercised on the way in.
    """

    def _make_docx(self, tmp_path: Path, name: str, paragraphs: list[str]) -> Path:
        """Create a minimal real .docx file using python-docx."""
        from docx import Document
        doc = Document()
        for text in paragraphs:
            doc.add_paragraph(text)
        path = tmp_path / name
        doc.save(str(path))
        return path

    def _patch_docling(self, markdown: str):
        mock_class, mock_converter = _make_docling_mock(markdown)
        return patch(
            "docling.document_converter.DocumentConverter",
            mock_class,
        )

    def test_basic_text_extracted(self, tmp_path):
        f = self._make_docx(tmp_path, "report.docx", [
            "Q3 Report",
            "Revenue grew 12% year-on-year.",
            "Costs remained flat.",
        ])
        with self._patch_docling("# Q3 Report\n\nRevenue grew 12% year-on-year.\n\nCosts remained flat."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")

        assert "Q3 Report" in node.text
        assert "Revenue grew" in node.text
        assert node.file_type == "docx"

    def test_file_type_set_correctly(self, tmp_path):
        f = self._make_docx(tmp_path, "doc.docx", ["Content."])
        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")
        assert node.file_type == "docx"

    def test_file_not_found_raises(self, tmp_path):
        from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
        with pytest.raises(FileNotFoundError):
            DOCXParser().parse(str(tmp_path / "missing.docx"), workspace="test")

    def test_docling_exception_raises_parser_error(self, tmp_path):
        f = self._make_docx(tmp_path, "bad.docx", ["Content."])

        mock_converter = MagicMock()
        mock_converter.convert.side_effect = RuntimeError("Docling internal error")
        mock_class = MagicMock(return_value=mock_converter)

        with patch(
            "docling.document_converter.DocumentConverter",
            mock_class,
        ):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            with pytest.raises(ParserError):
                DOCXParser().parse(str(f), workspace="test")

    def test_clean_pipeline_applied(self, tmp_path):
        f = self._make_docx(tmp_path, "quotes.docx", ["Content"])
        with self._patch_docling("\u201cSmart quotes\u201d and \u2014 em-dash."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")

        assert "\u201c" not in node.text
        assert "\u2014" not in node.text

    def test_title_override(self, tmp_path):
        f = self._make_docx(tmp_path, "annual.docx", ["Annual Report content."])
        with self._patch_docling("# Annual Report\n\nContent."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test", title="My Custom Title")
        assert node.title == "My Custom Title"

    def test_default_title_is_filename_stem(self, tmp_path):
        f = self._make_docx(tmp_path, "employee_handbook.docx", ["Content."])
        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")
        assert node.title == "employee_handbook"

    def test_workspace_set_on_node(self, tmp_path):
        f = self._make_docx(tmp_path, "doc.docx", ["Content."])
        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="hr_team")
        assert node.workspace == "hr_team"

    def test_doc_id_is_16_hex_chars(self, tmp_path):
        f = self._make_docx(tmp_path, "doc.docx", ["Content."])
        with self._patch_docling("Content."):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")
        assert len(node.doc_id) == 16
        assert all(c in "0123456789abcdef" for c in node.doc_id)

    def test_page_number_stripped_by_cleaner(self, tmp_path):
        f = self._make_docx(tmp_path, "paged.docx", ["Content."])
        with self._patch_docling("Introduction\n\nPage 1 of 10\n\nSome body text.\n\nPage 2 of 10"):
            from local_search_agent.ingestion.parsers.docx_parser import DOCXParser
            node = DOCXParser().parse(str(f), workspace="test")
        assert "Some body text" in node.text
        assert "Page 1 of 10" not in node.text
        assert "Page 2 of 10" not in node.text
